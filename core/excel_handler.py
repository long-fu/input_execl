"""Excel 文件读写处理"""
from pathlib import Path
from openpyxl import Workbook, load_workbook


class ExcelHandler:
    """管理 Excel 工作簿的读写"""

    def __init__(self, filepath: str | None = None):
        self._filepath: str | None = None
        if filepath and Path(filepath).exists():
            self._wb = load_workbook(filepath)
            self._filepath = filepath
        else:
            self._wb = Workbook()
            self._wb.active.title = "Sheet1"

    @property
    def filepath(self) -> str | None:
        return self._filepath

    @property
    def sheet(self):
        return self._wb.active

    @property
    def max_col(self) -> int:
        """返回当前 sheet 的数据最大列数（至少 1）"""
        mc = self.sheet.max_column
        return mc if mc else 1

    @property
    def max_row(self) -> int:
        """返回当前 sheet 的数据最大行数（至少 1）"""
        mr = self.sheet.max_row
        return mr if mr else 1

    def write_cell(self, col: int, row: int, value):
        """写入单元格，col/row 从 1 开始"""
        self.sheet.cell(row=row, column=col, value=value)

    def read_cell(self, col: int, row: int):
        """读取单元格值，返回字符串或 None"""
        cell = self.sheet.cell(row=row, column=col)
        return cell.value

    def get_matrix(self) -> list[list]:
        """获取整个数据区域的二维矩阵
        返回 list[list]，matrix[row][col]，索引从 0 开始
        """
        mr = self.max_row
        mc = self.max_col
        matrix = []
        for r in range(1, mr + 1):
            row_data = []
            for c in range(1, mc + 1):
                val = self.read_cell(c, r)
                row_data.append(val if val is not None else "")
            matrix.append(row_data)
        return matrix

    def save(self):
        """保存到当前文件，如未指定路径则调用 save_as"""
        if self._filepath:
            self._wb.save(self._filepath)
        else:
            raise ValueError("未指定文件路径，请使用 save_as")

    def save_as(self, filepath: str):
        """另存为新文件"""
        self._filepath = filepath
        self._wb.save(filepath)

    def new(self):
        """新建空白工作簿"""
        self._filepath = None
        self._wb = Workbook()
        self._wb.active.title = "Sheet1"
